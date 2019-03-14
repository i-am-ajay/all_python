__author__ = 'gaa8664'
import openpyxl


class ExcelProcessor:
    def __init__(self, file_path):
        self.__workbook = openpyxl.load_workbook(file_path)
        self.__head_map = {}
        self.__sequence_map = {}


    def set_workbook(self, file_path):
        self.__workbook = openpyxl.load_workbook(file_path)


    def get_workbook(self):
        return self.__workbook


    # returns list of names of all sheets in workbook
    def get_workbook_sheets(self):
        return self.__workbook.get_sheet_names()


    # returns name of active sheet
    def get_active_work_sheet_name(self):
        sheet = self.__workbook.get_active_sheet()
        return sheet.title


    # returns sheet from workbook, if name is not given then returns active workbook
    def get_sheet(self, sheet_name=''):
        sheet = None
        if sheet_name == '':
            sheet = self.get_workbook().get_active_sheet()
        else:
            sheet = self.get_workbook().get_sheet_by_name(sheet_name)
        return sheet


    # Get excel column heads
    def get_sheet_col_heads(self, sheet_name=''):
        count = 0
        sheet = self.get_sheet(sheet_name)
        for row in sheet:
            for col in row:
                col_val = sheet.cell(row=0, column=count).value
                self.head_map[1] = col_val
                self.sequence_map[col_val] = count
                count += 1
            break


