import openpyxl
import pprint

'''
    A program to split a string on the basis of given strings, like here string is split on basis of week days.
'''

string = 'Mon 11:00AM - 12:00PM, 2:00PM - 4:00PMTue 11:00AM - 12:00PM, 2:00PM - 4:00PMWed 11:00AM - 12:00PM, 2:00PM - 4:00PMThu 11:00AM - 12:00PM, 2:00PM - 4:00PMFri 11:00AM - 12:00PM, 2:00PM - 4:00PMSat 11:00AM - 12:00PM, 2:00PM - 4:00PM'

days_list = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']

days_index = {}

# Find index of each splitter string in the given string
def find_days_index(string):
    for day in days_list:
        days_index[day] = string.find(day)

# after splitting string, find the index of next immediate splitter so that relevant string could be extracted. Relevant string is between two splitters.
def index_of_next_day(string):
    nearest_index = 99999
    nearest_day = ''
    for day in days_list:
        index = string.find(day)
        # if any splitter exists in the given string then process the string and extract relevant data.
        if index != -1:
            if nearest_index > index:
                nearest_index = index
                nearest_day = day
    if nearest_day != '':
        return string.split(nearest_day)[0]
    else:
        return string


def dayswise_time_slots(string):
    day_timing_map = {}
    if string:
        find_days_index(string)
        for day in days_list:
            if string.find(day)>-1:
                sub_str = string.split(day)[1]
                #print(sub_str)
                sub_str = index_of_next_day(sub_str.strip())
                #print(day +sub_str)
                day_timing_map[day] = sub_str
    return day_timing_map


def insert_into_excel():
    col_timing_map = {'Mon':'K','Tue':'L','Wed':'M','Thu':'N','Fri':'O','Sat':'P','Sun':'Q'}
    excel = openpyxl.load_workbook('C:\\Users\\gaa8664\\Desktop\\for Slots.xlsx')
    sheet = excel.get_sheet_by_name('Sheet1')
    timing_sheet = excel.create_sheet(title="timing_sheet1")
    timing_sheet.append(['GAA']+days_list)
    for x in range(2,sheet.max_row):
        val = sheet['J'+str(x)]
        timing_map = dayswise_time_slots(val.value)
        '''first_col = 'A'+str(x)
        timing_sheet[first_col] = sheet[first_col].value'''
        if timing_map:
            for key in timing_map:
                sheet_index = col_timing_map[key]+str(x)
                print(sheet_index)
                print(sheet[sheet_index].value)
                if not sheet[sheet_index].value:
                    sheet[sheet_index] = ' '.join([key,timing_map[key]])

    excel.save('C:\\Users\\gaa8664\\Desktop\\for Slots.xlsx')

insert_into_excel()
