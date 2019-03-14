__author__ = 'gaa8664'
import openpyxl


'''program removes blank lines from work sheet and blank worksheet form workbook'''
def process_workbook(file):
    print('hello')
    workbook = openpyxl.load_workbook(file)
    for sheet_name in workbook.sheetnames:
        delete_blank_sheet(workbook, sheet_name)
    workbook.save(file)


def remove_blank_rows():
    pass


def delete_blank_sheet(workbook, sheet_name):
    sheet = workbook[sheet_name]
    if sheet.max_row <= 1:
        del workbook[sheet_name]


if __name__ == '__main__':
    process_workbook('D://comparison.xlsx')