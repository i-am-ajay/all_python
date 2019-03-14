__author__ = 'gaa8664'
import openpyxl
from salary import global_vars

def create_workbook():
    workbook = openpyxl.Workbook()
    return workbook


def create_sheet_with_headers(workbook, sheet_name, heads_description, file_path):
    sheet = workbook.create_sheet(sheet_name)
    count = 2
    head_map = {}
    for head in heads_description:
        head_map[head] = count
        head_desc = global_vars.FOX_PROD_SALHEADS[head]
        head_desc = head_desc.replace('9||','')
        fox_head_desc = 'Fox_'+ head_desc
        prod_head_desc = 'Prod_'+ head_desc
        sheet.cell(row=1, column=count).value = fox_head_desc
        count += 1
        sheet.cell(row=1, column=count).value = prod_head_desc
        count += 1
    sheet.cell(row=1, column=1).value = 'Emp Code'
    workbook.save(file_path)
    return sheet, head_map


def create_sheets_as_heads(workbook, heads_description, file_path):
    for head in heads_description:
        head_desc = global_vars.FOX_PROD_SALHEADS[head]
        head_desc = head_desc.replace('9||','')
        loc = head_desc.find('*')
        if loc > -1:
            head_list = head_desc.split('*')
            head_desc = head_list[0]
        sheet = workbook.create_sheet(head_desc.lower())
        fox_head_desc = 'Fox_'+ head_desc
        prod_head_desc = 'Prod_'+ head_desc
        sheet.cell(row=1, column=1).value = 'Emp Code'
        sheet.cell(row=1, column=2).value = fox_head_desc
        sheet.cell(row= 1, column=3).value =  prod_head_desc
    workbook.save(file_path)
    return sheet

