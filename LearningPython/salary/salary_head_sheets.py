# a program to divide excel sheets columns in seperate excel sheets.
import openpyxl

sal_workbook = openpyxl.Workbook()
salary_head_list = ['PR_AIDA','PR_IBASIC','PR_IBONUS','PR_ITS','PR_IDA','DAY_AMT','DAY_OFF','PR_IDIRTY','EL_AMT','PR_EL','PR_DELEC','PR_DEPF','PR_IHRA','HRS_AMT','PR_DIT','PR_DLWPVAL','PR_LWB','PR_DMESSAL','N_CARE','HRS','PR_DOTH_1','PR_IOTALW','P_CARE','PR_DPF','PR_ISP','PR_DOTHER','PR_LICAMT','PR_IUNIFOR','PR_DVOLPF','PR_IWASHIN']
# read given columns from the main workbook and create excel sheet
def create_excel_sheets(workbook, sheet_name, columns, head_name ):
    workbook = openpyxl.load_workbook(workbook)
    sheet = workbook.get_sheet_by_name(sheet_name)
    new_sheet = sal_workbook.create_sheet(title= head_name)
    new_sheet.cell(row=1,column=1).value="EmpCode"
    new_sheet.cell(row=1,column=2).value= head_name
    count = 0
    for rows in sheet:
        if count == 0:
            count += 1
            continue
        count += 1
        new_sheet.cell(row=count,column=1).value = rows[columns[0]].value
        new_sheet.cell(row=count, column=2).value = rows[columns[1]].value
    sal_workbook.save('D:/july/salary_workbook.xlsx')

def columns_map(filename, sheet_name):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name(sheet_name)
    for row in sheet.rows:
        count = 0
        for column in row:
            col_name = column.value
            if col_name == 'PR_NEWCODE':
                count += 1
                continue
            if col_name in salary_head_list:
                create_excel_sheets(filename, sheet_name,[9, count],col_name)
            count += 1
        break

if __name__ == '__main__':
    file = 'D:\\Software\\Software\\HR Module\\Salary\\Salary July 2017.xlsx'
    sheet_name = 'JUL17'
    columns_map(file, sheet_name)
