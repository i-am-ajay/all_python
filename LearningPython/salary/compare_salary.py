import salary
from salary.create_workbook import create_sheet_with_headers, create_sheets_as_heads

__author__ = 'gaa8664'
from salary.get_sheet import get_sheet
from salary import global_vars
from salary import prod_sal
from salary import create_maps
from salary.connection import Connection
from salary import fox_salary_map
from salary import prod_salary_map
import openpyxl


comparison_workbook = None
comp_sheet = None
head_map = None
row_count = 1
found = False


def initialize():
    global head_map
    global comp_sheet
    global comparison_workbook
    comparison_workbook = openpyxl.Workbook()
    comp_sheet, head_map = create_sheet_with_headers(comparison_workbook, 'comparison', global_vars.FOX_PROD_SALHEADS.keys(),
                        'D://comparison.xlsx')
    ''' comp_sheet = create_sheets_as_heads(comparison_workbook, global_vars.FOX_PROD_SALHEADS.keys(), 'D://comparison.xlsx')'''


def sal_comparison(file_name, sheet, year, month):
    emp_sal_map = {}
    fox_salary = {}
    count = 0
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.get_sheet_by_name(sheet)
    empcode_col = global_vars.HEAD_SEQUENCE['PR_NEWCODE']
    for row in sheet:
        global found_emp
        global row_count
        row_count = 1
        found_emp = False
        if found_emp == True:
            row_count += 1
        '''if count == 100:
            break'''
        fox_salary.clear()
        emp_sal_map.clear()
        # ignore first row of the sheet.
        if count == 0:
            count += 1
            continue
        empcode = row[empcode_col].value
        # employee salary result set
        emp_sal_resultset = prod_sal.get_employee_pro_sal(empcode, year, month)
        '''
        Process employee record from fox salary row and Prodigious Resultset to create two maps.
        1) Foxpro Map
        2) Prodigious Map
        Both map will have common keys for direct and simple comparison. A common value is set to represnt salary heads
        of both systems.
        '''
        if emp_sal_resultset:
            fox_salary = fox_salary_map.generate_fox_sal_map(row) # employee fox salary map
            #process result set to get salary head and value
            for val in emp_sal_resultset:
                emp_sal_map[val[2]] = val[0]
                #print("{1} -> {0},{2}".format(val[0],val[1], val[2]))
            prod_salary = prod_salary_map.generate_prod_sal_map(emp_sal_map) # employee prod salary map
            compare_emp_salary(empcode, fox_salary, prod_salary,row,count)
        #print(count)
        count += 1
    Connection.connection_close()
    Connection.close_cursor()
    comparison_workbook.save('D://comparison.xlsx')


def enter_val_comparision_sheet(head,fox_val, prod_val, empcode, row_count):
    '''sheet_name_work = get_sheet_name(head)
    if sheet_name:
        sheet = comparison_workbook.get_sheet_by_name(sheet_name_work.lower())
        sheet.cell(row=row_count, column=1).value = empcode
        sheet.cell(row = row_count, column=2).value = fox_val
        sheet.cell(row = row_count, column=3).value = prod_val'''
    comp_sheet.cell(row=row_count,column=1).value = empcode
    # map common head with fox heads
    fox_head = global_vars.COMMON_HEAD_FOX_HEAD_MAP[head]
    # get the column number where fox head will apper in excel.
    col_num = head_map.get(fox_head,None)
    if col_num:
        comp_sheet.cell(row=row_count, column=col_num).value = fox_val
        comp_sheet.cell(row=row_count, column=col_num+1).value = prod_val


def get_sheet_name(head):
    sheet_name = ''
    fox_head = global_vars.COMMON_HEAD_FOX_HEAD_MAP[head]
    prod_head = global_vars.FOX_PROD_SALHEADS.get(fox_head, None)
    if prod_head:
        sheet_name = prod_head.replace('9||','')
        loc = sheet_name.find('*')
        if loc > -1:
            str_array = sheet_name.split('*')
            sheet_name = str_array[0]
    return sheet_name


def compare_emp_salary(empcode, fox_salary, prod_salary, row,count):
    for head in fox_salary:
        prod_amount = prod_salary.get(head, 0.0)
        fox_amount = fox_salary[head]
        if head == 'Level':
            if fox_amount:
                try:
                    fox_amount = float(fox_amount.replace('Level-', ''))
                    if prod_amount == 'CONSOLIDATED SALARY' or prod_amount == 'FIXED STIPEND':
                        prod_amount = '0'
                except:
                    pass
            if fox_amount != prod_amount:

                enter_val_comparision_sheet(head, fox_amount, prod_amount,empcode, count)
        elif float(fox_amount) != float(prod_amount):
            print(empcode)
            enter_val_comparision_sheet(head, fox_amount, prod_amount,empcode, count)


if __name__ == '__main__':
    file_name = 'D:\Software\Software\HR Module\Salary\Salary April 2017.xlsx'
    sheet_name = 'PAYROLL'
    # create maps
    create_maps.head_mapping(file_name,'Map')
    create_maps.sequence_mapping(file_name, sheet_name)
    create_maps.common_head_fox_mapping()
    # initalize comparison sheet
    initialize()
    # compare salary
    sal_comparison(file_name, sheet_name, '2017', '04')

