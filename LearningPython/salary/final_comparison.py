import salary
from salary.create_workbook import create_sheet_with_headers, create_sheets_as_heads

__author__ = 'gaa8664'
from salary.get_sheet import get_sheet
from salary import global_vars
from salary import prod_sal
from salary import create_maps
from salary.connection import Connection
from salary import fox_salary_map
from salary import prod_salary_map
from salary import parseworkbook
from salary import emp_gross_sal
import openpyxl


comparison_workbook = None
comp_sheet = None
head_map = None
row_count = 1
found = False

def sal_comparison(file_name, sheet, month, year):
    # map specifying columns of comparison sheets.
    comparison_sheet_col = {1:'Emp Code', 2:'Foxpro Net', 3:'Prod Net',4:'Foxpro Deduction', 5:'Prod Deduction'}
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.get_sheet_by_name(sheet)

    comparison_workbook = openpyxl.Workbook()
    comp_sheet = comparison_workbook.create_sheet("Total Comparison")

    foxpro_empmaster_col = {}
    empcode_index = ''
    emp_deduct_index = ''
    emp_net_index = ''
    emp_depcode = ''

    count = 0

    for row in sheet:
        if count == 0:
            col_count = 0
            # find index of following columns in foxpro excel sheet.
            for col in row:
                if col.value == 'PR_NEWCODE':
                    empcode_index = col_count
                elif col.value == 'DEDUCT':
                    emp_deduct_index = col_count
                elif col.value == 'NET':
                    emp_net_index = col_count
                elif col.value == 'PR_DEPCODE':
                    print(col.value)
                    emp_depcode = col_count
                col_count += 1
            # create header in comparison sheet
            for index in range(1,6):
                comp_sheet.cell(row=1, column=index).value = comparison_sheet_col[index]
            count += 1
            continue
        if row[emp_depcode].value != 'OUT':
            empcode = row[empcode_index].value
            # employee salary result set
            emp_sal_resultset = emp_gross_sal.emp_gross_sal(empcode, month, year)
            '''
            Process employee record from fox salary row and Prodigious Resultset to create two maps.
            1) Foxpro Map
            2) Prodigious Map
            Both map will have common keys for direct and simple comparison. A common value is set to represnt salary heads
            of both systems.
            '''
            prod_net_amount = 0.0
            prod_deduction = 0.0
            if emp_sal_resultset:
                for record in emp_sal_resultset:
                    if record[1]==0:
                        prod_net_amount = record[2] # prodigious net
                    elif record[1] == 1:
                        prod_deduction = record[2] # prodigious deduction
            #compare_emp_salary(empcode, fox_salary, prod_salary,row,count)
            #print(count)
            fox_net_amount = row[emp_net_index].value
            fox_deduction = row[emp_deduct_index].value
            comp_sheet.cell(row=count, column=1).value = empcode
            comp_sheet.cell(row=count, column=2).value = fox_net_amount
            comp_sheet.cell(row=count, column=3).value = float(prod_net_amount) - float(prod_deduction)
            comp_sheet.cell(row=count, column=4).value = fox_deduction
            comp_sheet.cell(row=count, column=5).value = prod_deduction
            count += 1
        else:
            print(row[empcode_index].value)
    Connection.connection_close()
    Connection.close_cursor()
    # output file path
    comparison_workbook.save('E:\\Salary\Jan2019\\Salary total Comp_2.xlsx')

if __name__ == '__main__':
    print('hello')
    file_name = 'E:\\Salary\\Jan2019\\Salary January 2019.xlsx'
    sheet_name = 'Jan 2019'
    # compare salary
    sal_comparison(file_name, sheet_name, '01', '2019')


