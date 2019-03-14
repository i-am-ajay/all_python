__author__ = 'gaa8664'

import openpyxl

head_seq = {}
payrol_head_seq = {}


def create_sheet(fox_file_name,lwp_file_name, payrol_sheet_name, lwp_sheet_name, new_sheet_name):
    payrol_workbook = openpyxl.load_workbook(fox_file_name)
    lwp_workbook = openpyxl.load_workbook(lwp_file_name)

    payrol_sheet = payrol_workbook.get_sheet_by_name(payrol_sheet_name)
    lwp_sheet = lwp_workbook.get_sheet_by_name(lwp_sheet_name)

    new_sheet = lwp_workbook.create_sheet(new_sheet_name)
    payrol_sheet_header = payrol_sheet.rows[0]
    payrol_sheet_header_val = tuple((col.value for col in payrol_sheet_header))
    lwp_sheet_header = lwp_sheet.rows[0]
    count = 0
    for val in lwp_sheet_header:
        count += 1
        col_count = payrol_sheet_header_val.index(val.value)
        new_sheet.columns[count] = payrol_sheet.columns[col_count]

    lwp_workbook.save(lwp_file_name)


if __name__ == '__main__':
    fox_file_name = 'D:\\Software\\Software\\HR Module\\Salary\\Salary May 2017.xlsx'
    lwp_file_name = 'D:\\Software\\Software\\HR Module\\Salary\\LWP Example.xlsx'
    fox_sheet_name = 'May 2017'
    lwp_sheet_name = 'Sheet2'
    new_sheet_name = 'May'
    create_sheet(fox_file_name, lwp_file_name, fox_sheet_name, lwp_sheet_name, new_sheet_name)
