__author__ = 'gaa8664'
from salary.arrear_cal import connection
from salary.arrear_cal import Query
from salary.arrear_cal.Globals import month_dict
import openpyxl


con = connection.con_mysql
workbook = None
sheet = None
# stores col index of month header in excel sheet.
month_header_index = {}


def get_employee():
    global sheet
    cursor = con.cursor()
    cursor.execute(Query.GET_EMP_FULL_DETAILS)
    resultset = cursor.fetchall()
    row_count = 3
    for emp in resultset:
        sheet.cell(row = row_count, column = 1).value = emp[0]
        sheet.cell(row = row_count,column = 2).value = emp[1]
        sheet.cell(row = row_count, column = 3).value = emp[2]
        sheet.cell(row = row_count, column = 4).value = emp[3]
        sheet.cell(row = row_count, column = 5).value = emp[4]
        sheet.cell(row = row_count, column = 6).value = emp[5]
        get_employee_salary_data(emp[0],row_count)
        row_count += 1


def get_employee_salary_data(emp,row_count):
    global sheet
    cursor = con.cursor()
    cursor.execute(Query.EMP_SAL_ARREAR,(emp,))
    resultset = cursor.fetchall()
    for record in resultset:
        sal_month = record[3].month
        sal_year = record[3].year
        sal_period = month_dict[sal_month]+"_"+str(sal_year)
        # find the index of salary month in the excel sheet.
        col_index = month_header_index[sal_period]
        sheet.cell(row = row_count, column = col_index).value = record[2]
        sheet.cell(row=row_count,column = col_index+1).value = record[4]
        #emp_data.append([record[2],record[4],record[3]])



def create_excel_with_header():
    global workbook
    global sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet(title = "test_1")

    cursor = con.cursor()
    cursor.execute(Query.EMP_ARREAR_MONTH)
    recordset = cursor.fetchall()
    month_list = ['emp']
    index = 7
    sheet.cell(row=1,column=1).value = "Emp"
    sheet.cell(row=1,column=2).value = "Emp_name"
    sheet.cell(row=1, column=3).value = "Designation"
    sheet.cell(row=1, column=4).value = "Dept"
    sheet.cell(row=1, column=5).value = "Category"
    sheet.cell(row=1, column=6).value = "Status"
    for record in recordset:
        d = record[0]
        month_header = month_dict[d.month]+"_"+str(d.year)
        month_list.append(month_header)
        month_header_index[month_header] = index
        sheet.merge_cells(start_row = 1, start_column = index, end_row = 1, end_column = index+1)
        sheet.cell(row=1,column=index).value = month_header
        sheet.cell(row=2,column=index).value = "Amount"
        sheet.cell(row=2,column=index+1).value = "Minimum Wages"
        index += 2
    #sheet.append(month_list)
    print(month_list)



if __name__ == "__main__":
    #get_employee()
    create_excel_with_header()
    get_employee()
    workbook.save('test_file.xlsx')