__author__ = 'gaa8664'

import openpyxl

map_fox_to_prod_head = {}

def populate_head_dict(file, sheet):
    global map_fox_to_prod_head
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet)
    for record in sheet:
        head = record[0].value
        value = record[1].value
        if value :
            print('{}->{}'.format(head,value))
            map_fox_to_prod_head[head] = value

def change_fox_header(file, sheet):
    global map_fox_to_prod_head
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet)

    for col in range(1,sheet.max_column):
        val = sheet.cell(row=1, column = col).value
        try:
            cell_value = map_fox_to_prod_head[val]
            sheet.cell(row=1,column=col).value = cell_value
        except:
            print(val)
            continue
    workbook.save(file)

if __name__ == '__main__':
    populate_head_dict('E:\\Class B Salary_Dec_2018.xlsx','Map')
    change_fox_header('E:\\Class B Salary_Dec_2018.xlsx', 'register1')
