__author__ = 'gaa8664'

__author__ = 'gaa8664'
import openpyxl


workbook = openpyxl.load_workbook("C:\\Users\\gaa8664\\Desktop\\Employee Wise Templates Allocation (2)_1.xlsx")
fox_sheet = workbook.get_sheet_by_name("Sheet3")
prod_cat_sheet = workbook.get_sheet_by_name("Prod Cat")


def read_excel():
    col_title_map = {}
    col_title_seq = {}
    row_num = prod_cat_sheet.max_row
    col_num = prod_cat_sheet.max_column
    for row in range(1,row_num+1):
            if row == 1:
                for col in range(1, col_num+1):
                    col_title_map[col] = prod_cat_sheet.cell(row=row, column=col).value
                    col_title_seq[col_title_map[col]] = col
            else:
                empcode_col = col_title_seq['Employee']
                empcode = prod_cat_sheet.cell(row=row, column=empcode_col).value
                (foxcode,oldcode) = find_record(empcode)
                #print('{}:{} -> {}:{}'.format(empcode,cat, foxcode, fox_cat))
                if foxcode != None or foxcode != '':
                    fox_code_num = col_title_seq['Empcode']
                    fox_oldcode_num = col_title_seq['Oldcode']
                    #fox_name_num = col_title_seq['Namef']
                    #fox_cat_num = col_title_seq['Category']
                    prod_cat_sheet.cell(row=row, column=fox_code_num).value = foxcode
                    prod_cat_sheet.cell(row=row, column=fox_oldcode_num).value = oldcode
                    #prod_cat_sheet.cell(row=row, column=fox_cat_num).value = fox_cat
    workbook.save("C:\\Users\\gaa8664\\Desktop\\Employee Wise Templates Allocation (2)_2.xlsx")


# find employee in fox excel
def find_record(empcode):
    col_title_map = {}
    col_title_seq = {}
    row_num = fox_sheet.max_row
    col_num = fox_sheet.max_column
    for row in range(1,row_num+1):
        fox_empcode = ''
        fox_oldcode = ''
        category = ''
        fox_name = ''
        if row == 1:
            for col in range(1, col_num+1):
                col_title_map[col] = fox_sheet.cell(row=row, column=col).value
                col_title_seq[col_title_map[col]] = col
        else:
            empcode = str(empcode).upper()
            empcode_col_num = col_title_seq['PR_NEWCODE']
            fox_empcode = fox_sheet.cell(row =row, column =empcode_col_num).value
            fox_empcode = str(fox_empcode).upper()
            if empcode == fox_empcode:
                #cat_col_num = col_title_seq['Category']
                name_col_num = col_title_seq['PR_EMPCODE']
                #category = fox_sheet.cell(row=row, column=cat_col_num).value
                fox_oldcode = fox_sheet.cell(row=row, column=name_col_num).value
                fox_oldcode = str(fox_oldcode).upper()
                break
            fox_empcode = ''
    return (fox_empcode, fox_oldcode)

if __name__ == '__main__':
    read_excel()