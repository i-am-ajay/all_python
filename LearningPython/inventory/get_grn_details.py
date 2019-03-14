__author__ = 'gaa8664'

import openpyxl
import pymssql
from inventory import connection
from inventory import query

def process_grn_data():
    pass

def get_excel(path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']
    max_row = sheet.max_row
    max_col = sheet.max_column
    return workbook, sheet, max_row, max_col


def process_excel(sheet, max_row, max_col):
    for row in range(1,max_row+1):
        for col in range(1,max_col+1):
            data = sheet.cell(row=row,column=col)
            data1 = '{} -> {}'.format(col, data.value)
            print(data1)
        break
    count = 1
    for row in range(1, max_row+1):
        grn_number = sheet.cell(row=row, column=17).value
        if grn_number == None:
            item_code = sheet.cell(row=row, column=3).value
            grn_no = sheet.cell(row=row, column=12).value
            print('{}->{}'.format(item_code, grn_no))
            if grn_no is not None:
                tuple = find_grn_details(grn_no,item_code)
                print(tuple)
                if tuple is not None:
                    insert_grn_data(sheet, row, tuple)
                else:
                    recordset = get_opening_stock(item_code, grn_no, True)
                    insert_opening_stock(sheet, row, recordset)
            else:
                grn_batch = sheet.cell(row=row, column=9).value
                recordset = get_opening_stock(item_code, grn_batch, False)
                insert_opening_stock(sheet,row ,recordset)
            count += 1

def find_grn_details(grn, item):
    cursor = connection.get_connection().cursor()
    cursor.execute(query.FIND_GRN_DETAILS,(grn,item))
    #cursor.call_proc('sgrhsp_HMS_GRNITEMS_SearchDetails', ('', '', grn,'', '', '', item, '', '', '', '', '', ''))
    data = cursor.fetchall()
    if not data:
        cursor.close()
        return
    else:
        for row in data:
            grn_date = row[1]
            grn_no = row[7]
            grn_status = row[13]
            store_name = row[14]
            vendor_name = row[15]
            qty = row[20]
            rate = row[23]
            disc = row[24]
            tax = row[25]
            mrp = row[22]
            final_value = row[21]
            print('{}{}{}'.format(grn,item,(grn_date, grn_no,grn_status,store_name, vendor_name, qty, rate, disc, tax, mrp)))
    cursor.close()
    return (grn_date, grn_no,grn_status,store_name, vendor_name, qty, rate, disc, tax, mrp, final_value)


def get_opening_stock(item_code, batch_or_grn, is_grn):
    cursor = connection.get_proddb_connection().cursor(as_dict= True)
    if not is_grn:
        cursor.execute(query.OPENING_STOCK,(item_code, batch_or_grn))
        recordset = cursor.fetchone()
    else:
        cursor.execute(query.OPENING_STOCK_GRN,(item_code, batch_or_grn))
        recordset = cursor.fetchone();
    print(recordset)
    cursor.close()
    return recordset


def insert_grn_data(sheet, row,  tuple):
    sheet.cell(row= row,column=16).value = tuple[0]
    sheet.cell(row= row, column=17).value = tuple[1]
    sheet.cell(row= row, column=18).value = tuple[2]
    sheet.cell(row= row, column=19).value = tuple[3]
    sheet.cell(row= row, column=20).value = tuple[4]
    sheet.cell(row= row, column=23).value = tuple[5]
    sheet.cell(row= row, column=24).value = tuple[6]
    sheet.cell(row= row, column=25).value = tuple[7]
    sheet.cell(row= row, column=27).value = tuple[8]
    sheet.cell(row=row, column=31).value = tuple[10]
    sheet.cell(row= row, column=33).value = tuple[9]


def insert_opening_stock(sheet, row, recordset):
    if recordset:
        sheet.cell(row=row, column=16).value = recordset['GRNDate']
        sheet.cell(row=row, column=17).value = recordset['GRNNo']
        sheet.cell(row=row, column=3).value = recordset['ITEM']
        sheet.cell(row=row, column=33).value = recordset['MRP']
        sheet.cell(row=row, column=31).value = recordset['TotalValue']
        sheet.cell(row=row, column=9).value = recordset['BATCHNO']
        sheet.cell(row=row, column=1).value = recordset['Store']
        sheet.cell(row=row, column=23).value = recordset['GRNQTY']
        sheet.cell(row=row, column=24).value = recordset['ItemRate']


if __name__== '__main__':
    excel = get_excel('C:\\Users\\gaa8664\Desktop\\MRP.xlsx')
    workbook = excel[0]
    sheet = excel[1]
    max_row = excel[2]
    max_col = excel[3]
    process_excel(sheet, max_row, max_col)
    workbook.save('C:\\Users\\gaa8664\Desktop\\MRP.xlsx')
    #find_grn_details('SGRH-GRN/5360/2008-2009', 'COMACCDESPC1')

    connection.close_connection()
    connection.close_proddb_connection()





