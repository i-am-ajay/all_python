__author__ = 'gaa8664'
import openpyxl


def get_new_code(sheet, old_code):
    new_code = None
    for row in range(sheet.max_row):
        if row <= 1:
            continue
        code = sheet.cell(row=row, column=2)
        if old_code == code.value:
            try:
                new_code = sheet.cell(row=row, column=3)
                new_code = new_code.value

                break
            except Exception as ex:
                new_code = ''
    return new_code


def update_lic():
    code_file = openpyxl.load_workbook("D:\Project\excel\Employee Wise Templates Allocation.xlsx")
    lic_file = openpyxl.load_workbook("D:\Project\excel\LIC Detail.xlsx",read_only=False)
    code_sheet = code_file.get_sheet_by_name("Template Allocation")
    lic_sheet = lic_file.get_sheet_by_name("LIC Detail")

    for row in range(lic_sheet.max_row):
        if row <= 1:
            continue
        old_code = lic_sheet.cell(row=row, column=1)
        new_code = lic_sheet.cell(row=row, column=2)
        val = get_new_code(code_sheet, old_code.value)
        new_code.value = val
    lic_file.save("D:\Project\excel\LIC Detail1.xlsx")

if __name__ == '__main__':
    update_lic()







