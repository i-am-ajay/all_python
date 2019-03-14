__author__ = 'gaa8664'
import openpyxl


def get_stores_in_row():
    file_name = 'D:\\Medicine.xlsx'
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.get_sheet_by_name('Sheet1')
    count = 0
    col_num = sheet.max_column
    store_name = ''
    for row in sheet:
        if count == 0:
            count += 1
            continue
        if row[0].value == 'Store Name':
            store_name = row[1].value
            count += 1
            continue
        else:
            sheet.cell(row=count+1, column= col_num+1).value = store_name
            count += 1
    workbook.save(file_name)


if __name__ == '__main__':
    get_stores_in_row()