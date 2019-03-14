import openpyxl
import re # regular expression
import pprint

#global variables
master_dict = {}
pattren = re.compile("\{\w*\}")


# create comment code and desc dictionary, using comment master from excel sheet.
def comment_master_dict(master_sheet):
    for record in master_sheet:
       master_dict[str(record[1].value)] = record[2].value
    pprint.pprint(master_dict)

'''
    * Using comment_master we create a dictionary of comments. Method: comment_master_dict
    * Iterate through all the records of the report, get comment column.
    * Using comment code get the comment definition for comment_dict.
'''
def process_excel(master_sheet, report_sheet):
    comment_master_dict(master_sheet)
    count = 1
    last_col = report_sheet.max_column
    # Iterate report records
    for row in report_sheet:
        if count == 0:
            count += 1
            continue
        comment_code = row[last_col-1].value
        print(comment_code)
        if comment_code:
            write_comment = False
            comment = ''
            code = ''
            code_found = False
            # iterate through each single character and process it as per conditions.
            for c in comment_code:
                if c is '{':
                    code_found = True
                    write_comment = True
                else:
                    if code_found and c is not '}':
                        code += c
                    elif code_found and c is '}':
                        print(code)
                        print(master_dict.get(code))
                        code_des = master_dict.get(str(code))
                        comment +='{'+code_des+'}'
                        code=''
                        code_found = False
                    else:
                        comment += c
            if write_comment:
                report_sheet.cell(row=count, column=last_col+1).value = comment
            else:
                report_sheet.cell(row=count, column=last_col+1).value = comment_code
        count += 1

    report_data.save('E:\Project\Report\micro_main_1.xlsx')


if __name__ == '__main__':
    # comment excels
    comment_master = openpyxl.load_workbook('E:\Project\Report\micro2.xlsx')
    master_sheet = comment_master['Sheet1']

    # report workbook
    report_data = openpyxl.load_workbook('E:\Project\Report\micro_main.xlsx')
    report_sheet = report_data['Sheet1']

    process_excel(master_sheet, report_sheet)

# Regular expression based code.
''' code_list = pattren.finditer(comment_code) # list of substrings those matches our pattern, here comment codes.
            for code in code_list:
                code = code.group()[1:-1]
                comment+='{'+master_dict.get(code)+'}'''



