import openpyxl

workbook = None
worksheet = None
new_workbook = None
new_worksheet = None

patient_demo = {
    'SampleID' : '',
    'RegnNo' : '',
    'FirstName': '',
    'LastName': '',
    'PatientLocation': '',
    'PatientAge': '',
    'PatientSex': '',
}
test_info = {
    'Test_code':'',
    'Test_name':'',
}


def read_labdata(workbook, sheet):
    workbook = openpyxl.load_workbook(workbook)
    worksheet = workbook[sheet]
    for record in worksheet:
        if(patient_demo['sample_id'] == '' or patient_demo['sample_id'] != record[0]):
            patient_demo['sample_id'] = record[0]
        for col in record:
            print(col.value)
        break


def check_record(record):
    pass


new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.create_sheet()

if __name__ == '__main__':
    read_labdata('E:\Project\Report\Micro.xlsx','Sheet1')
